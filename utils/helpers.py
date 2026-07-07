# =============================================================================
# REPORTE CONSOLIDADO DE VISITAS - VERSIÓN MEJORADA
# =============================================================================

import io
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from typing import Optional, List, Tuple, Dict, Any

# =============================================================================
# CONFIGURACIÓN DE ESTILOS Y CONSTANTES
# =============================================================================

COLORES_REPORTE = {
    "primario": "#1B3A5C",
    "secundario": "#2E5A88",
    "exito": "#1E7E34",
    "exito_fondo": "#EAF7EE",
    "advertencia": "#B45309",
    "advertencia_fondo": "#FEF6E7",
    "peligro": "#C8102E",
    "peligro_fondo": "#FDECEC",
    "gris": "#D9DEE6",
    "gris_texto": "#55606E",
    "fondo": "#F8FAFC",
}

RESULTADO_VISITA_MAP = {
    "1. Cliente con actividad laboral y/o económica vigente": {
        "icono": "✔", 
        "color": COLORES_REPORTE["exito"],
        "fondo": COLORES_REPORTE["exito_fondo"],
        "abrev": "Activo"
    },
    "2. Cliente con situación desmejorada": {
        "icono": "⚠", 
        "color": COLORES_REPORTE["advertencia"],
        "fondo": COLORES_REPORTE["advertencia_fondo"],
        "abrev": "Desmejorado"
    },
    "3. Cliente ya no labora y/o no realiza la actividad económica": {
        "icono": "✖", 
        "color": COLORES_REPORTE["peligro"],
        "fondo": COLORES_REPORTE["peligro_fondo"],
        "abrev": "No labora"
    },
    "4. Cliente no ubicado": {
        "icono": "✖", 
        "color": COLORES_REPORTE["peligro"],
        "fondo": COLORES_REPORTE["peligro_fondo"],
        "abrev": "No ubicado"
    },
}

# =============================================================================
# FUNCIONES DE FILTRADO Y PROCESAMIENTO
# =============================================================================

def filtrar_historial(
    df: pd.DataFrame,
    agencia: Optional[str] = None,
    fecha_desde: Optional[datetime] = None,
    fecha_hasta: Optional[datetime] = None,
    analista: Optional[str] = None,
    resultado: Optional[str] = None,
) -> pd.DataFrame:
    """
    Aplica filtros dinámicos al historial de visitas.
    
    Args:
        df: DataFrame con el historial
        agencia: Filtrar por agencia
        fecha_desde: Fecha de inicio
        fecha_hasta: Fecha de fin
        analista: Filtrar por auditor
        resultado: Filtrar por resultado de visita
    
    Returns:
        DataFrame filtrado
    """
    if df.empty:
        return df
    
    df_filtrado = df.copy()
    
    # Filtro por agencia
    if agencia and "Agencia" in df_filtrado.columns:
        df_filtrado = df_filtrado[
            df_filtrado["Agencia"].astype(str).str.strip() == str(agencia).strip()
        ]
    
    # Filtro por fecha
    if "Fecha" in df_filtrado.columns:
        # Convertir fechas al formato adecuado
        try:
            df_filtrado["Fecha_dt"] = pd.to_datetime(
                df_filtrado["Fecha"], format="%d/%m/%Y", errors="coerce"
            )
            
            if fecha_desde:
                df_filtrado = df_filtrado[df_filtrado["Fecha_dt"] >= fecha_desde]
            if fecha_hasta:
                df_filtrado = df_filtrado[df_filtrado["Fecha_dt"] <= fecha_hasta]
        except Exception:
            pass
    
    # Filtro por analista
    if analista and "Usuario_Auditor" in df_filtrado.columns:
        df_filtrado = df_filtrado[
            df_filtrado["Usuario_Auditor"].astype(str).str.contains(analista, case=False, na=False)
        ]
    
    # Filtro por resultado
    if resultado and "ClienteVisitado" in df_filtrado.columns:
        df_filtrado = df_filtrado[
            df_filtrado["ClienteVisitado"].astype(str).str.contains(resultado, case=False, na=False)
        ]
    
    # Eliminar columna auxiliar si existe
    if "Fecha_dt" in df_filtrado.columns:
        df_filtrado = df_filtrado.drop(columns=["Fecha_dt"])
    
    return df_filtrado


def obtener_opciones_filtro(df: pd.DataFrame, columna: str) -> List[str]:
    """
    Obtiene las opciones únicas para un filtro.
    
    Args:
        df: DataFrame del historial
        columna: Nombre de la columna
    
    Returns:
        Lista de opciones únicas
    """
    if df.empty or columna not in df.columns:
        return []
    
    valores = df[columna].dropna().astype(str).str.strip().unique()
    return sorted([v for v in valores if v and v != "nan"])


def calcular_indicadores(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calcula los indicadores generales a partir de los datos filtrados.
    
    Args:
        df: DataFrame con los datos filtrados
    
    Returns:
        Diccionario con los indicadores
    """
    total_visitas = len(df)
    
    # Visitas verificadas = aquellas que tienen un resultado registrado
    if "ClienteVisitado" in df.columns:
        verificadas = df[
            df["ClienteVisitado"].astype(str).str.strip().notna() &
            (df["ClienteVisitado"].astype(str).str.strip() != "") &
            (df["ClienteVisitado"].astype(str).str.strip() != "nan")
        ]
        total_verificadas = len(verificadas)
    else:
        total_verificadas = total_visitas
    
    pendientes = total_visitas - total_verificadas
    cumplimiento = (total_verificadas / total_visitas * 100) if total_visitas > 0 else 0
    
    # Distribución por resultado
    distribucion = {}
    if "ClienteVisitado" in df.columns:
        for resultado, count in df["ClienteVisitado"].value_counts().items():
            resultado_str = str(resultado).strip()
            if resultado_str and resultado_str != "nan":
                distribucion[resultado_str] = {
                    "cantidad": count,
                    "porcentaje": (count / total_visitas * 100) if total_visitas > 0 else 0
                }
    
    return {
        "total": total_visitas,
        "verificadas": total_verificadas,
        "pendientes": pendientes,
        "cumplimiento": cumplimiento,
        "distribucion": distribucion,
    }


def preparar_detalle_visitas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara el detalle de visitas para el reporte.
    
    Args:
        df: DataFrame con los datos filtrados
    
    Returns:
        DataFrame con las columnas formateadas
    """
    if df.empty:
        return pd.DataFrame()
    
    columnas_detalle = [
        "Cuenta", "NumeroOperacion", "Cliente", "Modulo",
        "AnalistaVigente", "AnalistaEvaluador", "Resultado", "ClienteVisitado", "Fecha"
    ]
    
    # Mapeo de columnas
    mapeo = {
        "BCCTA": "Cuenta",
        "BCOPER": "NumeroOperacion",
        "CLIENTE": "Cliente",
        "MODULO": "Modulo",
        "ANALISTA": "AnalistaVigente",
        "ANALISTA_EVAL": "AnalistaEvaluador",
        "ClienteVisitado": "ClienteVisitado",
        "Fecha": "Fecha",
    }
    
    # Renombrar columnas
    df_detalle = df.copy()
    for col_orig, col_dest in mapeo.items():
        if col_orig in df_detalle.columns:
            df_detalle[col_dest] = df_detalle[col_orig]
    
    # Determinar resultado de la visita (SI/NO)
    df_detalle["Resultado"] = df_detalle["ClienteVisitado"].apply(
        lambda x: "SI" if pd.notna(x) and str(x).strip() and str(x).strip() != "nan" else "NO"
    )
    
    # Seleccionar columnas
    cols_existentes = [c for c in columnas_detalle if c in df_detalle.columns]
    df_detalle = df_detalle[cols_existentes].copy()
    
    # Limpiar valores
    for col in df_detalle.columns:
        df_detalle[col] = df_detalle[col].astype(str).str.strip()
        df_detalle[col] = df_detalle[col].replace(["nan", "None", ""], "-")
    
    return df_detalle


# =============================================================================
# GENERACIÓN DE REPORTE EXCEL (FORMATO MEJORADO)
# =============================================================================

def generar_reporte_consolidado_excel(
    df_historial: pd.DataFrame,
    agencia: Optional[str] = None,
    fecha_desde: Optional[datetime] = None,
    fecha_hasta: Optional[datetime] = None,
    analista: Optional[str] = None,
    resultado: Optional[str] = None,
    auditor_responsable: str = "",
) -> bytes:
    """
    Genera un reporte consolidado en Excel con formato similar a la imagen.
    
    Args:
        df_historial: DataFrame con el historial completo
        agencia: Filtro por agencia
        fecha_desde: Fecha de inicio
        fecha_hasta: Fecha de fin
        analista: Filtro por auditor
        resultado: Filtro por resultado
        auditor_responsable: Nombre del auditor que genera el reporte
    
    Returns:
        Bytes del archivo Excel
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    
    # Filtrar datos
    df_filtrado = filtrar_historial(df_historial, agencia, fecha_desde, fecha_hasta, analista, resultado)
    
    if df_filtrado.empty:
        # Crear Excel con mensaje de "Sin datos"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = "REPORTE CONSOLIDADO DE VISITAS"
        ws["A3"] = "No se encontraron registros con los filtros seleccionados."
        for col in "ABCDE":
            ws.column_dimensions[col].width = 20
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Calcular indicadores
    indicadores = calcular_indicadores(df_filtrado)
    df_detalle = preparar_detalle_visitas(df_filtrado)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    
    # --- ESTILOS ---
    # Colores
    COLOR_PRIMARIO = "1B3A5C"
    COLOR_FONDO = "F8FAFC"
    COLOR_GRIS = "D9DEE6"
    COLOR_EXITO = "1E7E34"
    COLOR_EXITO_FONDO = "EAF7EE"
    COLOR_ADVERTENCIA = "B45309"
    COLOR_ADVERTENCIA_FONDO = "FEF6E7"
    COLOR_PELIGRO = "C8102E"
    COLOR_PELIGRO_FONDO = "FDECEC"
    
    # Fuentes
    fuente_titulo = Font(name="Calibri", size=18, bold=True, color="1B3A5C")
    fuente_subtitulo = Font(name="Calibri", size=11, bold=True, color="1B3A5C")
    fuente_normal = Font(name="Calibri", size=10)
    fuente_negrita = Font(name="Calibri", size=10, bold=True)
    fuente_encabezado = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    fuente_resultado = Font(name="Calibri", size=10, bold=True)
    
    # Bordes
    borde_fino = Side(style="thin", color=COLOR_GRIS)
    borde_celda = Border(top=borde_fino, bottom=borde_fino, left=borde_fino, right=borde_fino)
    
    # Alineación
    alinear_centro = Alignment(horizontal="center", vertical="center")
    alinear_izquierda = Alignment(horizontal="left", vertical="center")
    
    # --- FUNCIONES AUXILIARES ---
    def set_cell(ws, row, col, value, style=None, alineacion=None, borde=None):
        """Establece una celda con formato."""
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = col
        cell = ws[f"{col_letter}{row}"]
        cell.value = value
        if style:
            cell.font = style
        if alineacion:
            cell.alignment = alineacion
        if borde:
            cell.border = borde
    
    def merge_cells(ws, start_row, start_col, end_row, end_col):
        """Combina celdas."""
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f"{start_letter}{start_row}:{end_letter}{end_row}")
    
    # --- ENCABEZADO ---
    fila_actual = 1
    
    # Título principal
    merge_cells(ws, fila_actual, 1, fila_actual, 8)
    set_cell(ws, fila_actual, 1, "REPORTE CONSOLIDADO DE VISITAS", 
             style=fuente_titulo, alineacion=alinear_centro)
    fila_actual += 2
    
    # Información del reporte (dos columnas)
    datos_reporte = [
        ("Agencia:", agencia or "Todas"),
        ("Auditor responsable:", auditor_responsable or st.session_state.get("usuario", "")),
        ("Fecha del reporte:", datetime.now().strftime("%d/%m/%Y")),
        ("Total de visitas:", indicadores["total"]),
    ]
    
    for i, (label, value) in enumerate(datos_reporte):
        col_label = 1 if i < 2 else 5
        col_value = 2 if i < 2 else 6
        row_offset = 2 if i < 2 else 0
        
        row = fila_actual + (i % 2)
        set_cell(ws, row, col_label, label, style=fuente_subtitulo, alineacion=alinear_izquierda)
        set_cell(ws, row, col_value, value, style=fuente_normal, alineacion=alinear_izquierda)
    
    fila_actual += 2
    
    # --- INDICADORES EN TARJETAS ---
    # Fila de indicadores
    indicadores_data = [
        ("TOTAL DE VISITAS", indicadores["total"], COLOR_PRIMARIO, "FFFFFF"),
        ("VISITAS VERIFICADAS", indicadores["verificadas"], COLOR_EXITO, "FFFFFF"),
        ("VISITAS PENDIENTES", indicadores["pendientes"], COLOR_PELIGRO, "FFFFFF"),
        ("% DE CUMPLIMIENTO", f"{indicadores['cumplimiento']:.1f}%", COLOR_ADVERTENCIA, "FFFFFF"),
    ]
    
    for idx, (label, value, bg_color, txt_color) in enumerate(indicadores_data):
        col = idx * 2 + 1
        col_end = col + 1
        
        # Celda combinada para la tarjeta
        merge_cells(ws, fila_actual, col, fila_actual, col_end)
        merge_cells(ws, fila_actual + 1, col, fila_actual + 1, col_end)
        
        # Valor grande
        cell = ws.cell(row=fila_actual, column=col)
        cell.value = value
        cell.font = Font(name="Calibri", size=22, bold=True, color=txt_color)
        cell.alignment = alinear_centro
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.border = borde_celda
        
        # Etiqueta
        cell2 = ws.cell(row=fila_actual + 1, column=col)
        cell2.value = label
        cell2.font = Font(name="Calibri", size=9, bold=True, color=txt_color)
        cell2.alignment = alinear_centro
        cell2.fill = PatternFill("solid", fgColor=bg_color)
        cell2.border = borde_celda
    
    fila_actual += 3
    
    # --- DISTRIBUCIÓN DE RESULTADOS ---
    # Título de la sección
    merge_cells(ws, fila_actual, 1, fila_actual, 8)
    set_cell(ws, fila_actual, 1, "DISTRIBUCIÓN DE RESULTADOS", 
             style=fuente_subtitulo, alineacion=alinear_izquierda)
    fila_actual += 1
    
    # Encabezados de la tabla
    headers = ["Resultado de la visita", "Cantidad", "Porcentaje", "", "Distribución"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila_actual, column=i)
        cell.value = header
        cell.font = fuente_encabezado
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        cell.alignment = alinear_centro
        cell.border = borde_celda
    
    fila_actual += 1
    
    # Datos de distribución
    distribucion = indicadores["distribucion"]
    
    # Ordenar según las opciones predefinidas
    orden_resultados = [
        "1. Cliente con actividad laboral y/o económica vigente",
        "2. Cliente con situación desmejorada",
        "3. Cliente ya no labora y/o no realiza la actividad económica",
        "4. Cliente no ubicado",
    ]
    
    fila_inicio_datos = fila_actual
    for resultado_text in orden_resultados:
        if resultado_text in distribucion:
            datos = distribucion[resultado_text]
            set_cell(ws, fila_actual, 1, resultado_text, style=fuente_normal, alineacion=alinear_izquierda)
            set_cell(ws, fila_actual, 2, datos["cantidad"], style=fuente_normal, alineacion=alinear_centro)
            set_cell(ws, fila_actual, 3, f"{datos['porcentaje']:.1f}%", style=fuente_normal, alineacion=alinear_centro)
            # Columna vacía (separador)
            set_cell(ws, fila_actual, 4, "", style=fuente_normal, alineacion=alinear_centro)
            # Distribución visual (barra de progreso simple)
            barra = "█" * int(datos["porcentaje"] / 5) if datos["porcentaje"] > 0 else "▁"
            set_cell(ws, fila_actual, 5, f"{barra} {datos['porcentaje']:.1f}%", 
                     style=fuente_normal, alineacion=alinear_izquierda)
            
            # Color de fondo según el resultado
            for col in range(1, 6):
                cell = ws.cell(row=fila_actual, column=col)
                if "actividad laboral" in resultado_text.lower():
                    cell.fill = PatternFill("solid", fgColor=COLOR_EXITO_FONDO)
                elif "desmejorada" in resultado_text.lower():
                    cell.fill = PatternFill("solid", fgColor=COLOR_ADVERTENCIA_FONDO)
                elif "no labora" in resultado_text.lower() or "no ubicado" in resultado_text.lower():
                    cell.fill = PatternFill("solid", fgColor=COLOR_PELIGRO_FONDO)
                cell.border = borde_celda
            
            fila_actual += 1
    
    # Fila de total
    set_cell(ws, fila_actual, 1, "TOTAL", style=fuente_negrita, alineacion=alinear_izquierda)
    set_cell(ws, fila_actual, 2, indicadores["total"], style=fuente_negrita, alineacion=alinear_centro)
    set_cell(ws, fila_actual, 3, "100%", style=fuente_negrita, alineacion=alinear_centro)
    for col in range(1, 6):
        cell = ws.cell(row=fila_actual, column=col)
        cell.border = borde_celda
        cell.font = fuente_negrita
    
    fila_actual += 2
    
    # --- DETALLE DE VISITAS ---
    merge_cells(ws, fila_actual, 1, fila_actual, 9)
    set_cell(ws, fila_actual, 1, "DETALLE DE VISITAS", 
             style=fuente_subtitulo, alineacion=alinear_izquierda)
    fila_actual += 1
    
    # Encabezados del detalle
    headers_detalle = ["N°", "Cuenta Cliente", "Número de Operación", "Nombre del Cliente", 
                       "Módulo", "Analista Vigente", "Analista Evaluador", "Resultado", "Cliente visitado"]
    
    for i, header in enumerate(headers_detalle, start=1):
        cell = ws.cell(row=fila_actual, column=i)
        cell.value = header
        cell.font = fuente_encabezado
        cell.fill = PatternFill("solid", fgColor=COLOR_PRIMARIO)
        cell.alignment = alinear_centro
        cell.border = borde_celda
    
    fila_actual += 1
    
    # Datos del detalle
    if not df_detalle.empty:
        for idx, row in df_detalle.iterrows():
            # Número secuencial
            num_fila = fila_actual - fila_inicio_datos
            set_cell(ws, fila_actual, 1, num_fila, style=fuente_normal, alineacion=alinear_centro)
            
            # Datos del cliente
            columnas_detalle = ["Cuenta", "NumeroOperacion", "Cliente", "Modulo", 
                               "AnalistaVigente", "AnalistaEvaluador", "Resultado", "ClienteVisitado"]
            
            for j, col in enumerate(columnas_detalle, start=2):
                valor = row.get(col, "-")
                set_cell(ws, fila_actual, j, valor, style=fuente_normal, 
                        alineacion=alinear_izquierda if j != 2 and j != 8 else alinear_centro)
            
            # Aplicar bordes
            for col in range(1, 10):
                cell = ws.cell(row=fila_actual, column=col)
                cell.border = borde_celda
                # Color de fondo para "Resultado"
                if col == 8:
                    if str(row.get("Resultado", "")).upper() == "SI":
                        cell.fill = PatternFill("solid", fgColor=COLOR_EXITO_FONDO)
                        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_EXITO)
                    else:
                        cell.fill = PatternFill("solid", fgColor=COLOR_PELIGRO_FONDO)
                        cell.font = Font(name="Calibri", size=10, bold=True, color=COLOR_PELIGRO)
            
            fila_actual += 1
    
    # --- PIE DE PÁGINA ---
    fila_actual += 1
    
    # Total de registros
    merge_cells(ws, fila_actual, 1, fila_actual, 3)
    set_cell(ws, fila_actual, 1, f"Total de registros: {indicadores['total']}", 
             style=fuente_subtitulo, alineacion=alinear_izquierda)
    
    fila_actual += 1
    
    # Fecha de generación
    merge_cells(ws, fila_actual, 1, fila_actual, 3)
    set_cell(ws, fila_actual, 1, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}", 
             style=fuente_normal, alineacion=alinear_izquierda)
    
    fila_actual += 1
    
    # Generado por
    merge_cells(ws, fila_actual, 1, fila_actual, 3)
    set_cell(ws, fila_actual, 1, f"Generado por: {auditor_responsable or st.session_state.get('usuario', '')}", 
             style=fuente_normal, alineacion=alinear_izquierda)
    
    # --- AJUSTE DE ANCHOS ---
    anchos = {
        "A": 5,   # N°
        "B": 16,  # Cuenta Cliente
        "C": 18,  # Número de Operación
        "D": 30,  # Nombre del Cliente
        "E": 25,  # Módulo
        "F": 18,  # Analista Vigente
        "G": 18,  # Analista Evaluador
        "H": 12,  # Resultado
        "I": 32,  # Cliente visitado
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    # Ajustar filas
    ws.row_dimensions[1].height = 30
    
    # --- CONFIGURACIÓN DE PÁGINA ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Margenes
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.7, bottom=0.7,
        header=0.3, footer=0.3
    )
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# =============================================================================
# FUNCIÓN PARA MOSTRAR FILTROS EN STREAMLIT
# =============================================================================

def mostrar_filtros_reporte(df_historial: pd.DataFrame) -> Tuple[Optional[str], Optional[datetime], Optional[datetime], Optional[str], Optional[str]]:
    """
    Muestra los filtros para el reporte consolidado en Streamlit.
    
    Returns:
        Tupla con (agencia, fecha_desde, fecha_hasta, analista, resultado)
    """
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Filtro por agencia
        agencias = obtener_opciones_filtro(df_historial, "Agencia")
        agencia_opciones = ["Todas"] + agencias
        agencia_seleccionada = st.selectbox("Agencia", agencia_opciones, key="filtro_agencia")
        agencia = None if agencia_seleccionada == "Todas" else agencia_seleccionada
    
    with col2:
        # Filtro por analista
        analistas = obtener_opciones_filtro(df_historial, "Usuario_Auditor")
        analista_opciones = ["Todos"] + analistas
        analista_seleccionado = st.selectbox("Auditor", analista_opciones, key="filtro_analista")
        analista = None if analista_seleccionado == "Todos" else analista_seleccionado
    
    with col3:
        # Filtro por resultado
        resultados = obtener_opciones_filtro(df_historial, "ClienteVisitado")
        resultado_opciones = ["Todos"] + resultados
        resultado_seleccionado = st.selectbox("Resultado de visita", resultado_opciones, key="filtro_resultado")
        resultado = None if resultado_seleccionado == "Todos" else resultado_seleccionado
    
    col4, col5 = st.columns(2)
    
    with col4:
        fecha_desde = st.date_input("Fecha desde", value=datetime.now() - timedelta(days=30), key="filtro_fecha_desde")
    
    with col5:
        fecha_hasta = st.date_input("Fecha hasta", value=datetime.now(), key="filtro_fecha_hasta")
    
    return agencia, fecha_desde, fecha_hasta, analista, resultado


# =============================================================================
# FUNCIÓN PARA MOSTRAR EL DASHBOARD DEL REPORTE
# =============================================================================

def mostrar_dashboard_reporte(df_historial: pd.DataFrame):
    """
    Muestra un dashboard interactivo del reporte consolidado en Streamlit.
    """
    st.markdown("""
    <style>
    .metric-card {
        background: #FFFFFF;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-left: 4px solid #1B3A5C;
        margin-bottom: 10px;
    }
    .metric-value {
        font-size: 28px;
        font-weight: bold;
        color: #1B3A5C;
    }
    .metric-label {
        font-size: 12px;
        color: #55606E;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Filtros
    agencia, fecha_desde, fecha_hasta, analista, resultado = mostrar_filtros_reporte(df_historial)
    
    # Aplicar filtros
    df_filtrado = filtrar_historial(df_historial, agencia, fecha_desde, fecha_hasta, analista, resultado)
    
    if df_filtrado.empty:
        st.warning("No se encontraron registros con los filtros seleccionados.")
        return
    
    # Calcular indicadores
    indicadores = calcular_indicadores(df_filtrado)
    
    # --- MÉTRICAS ---
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            label="📊 Total de Visitas",
            value=indicadores["total"],
            delta=None,
        )
    
    with col2:
        st.metric(
            label="✅ Verificadas",
            value=indicadores["verificadas"],
            delta=f"{indicadores['cumplimiento']:.1f}%",
            delta_color="normal",
        )
    
    with col3:
        st.metric(
            label="⏳ Pendientes",
            value=indicadores["pendientes"],
            delta=None,
        )
    
    with col4:
        st.metric(
            label="🎯 % Cumplimiento",
            value=f"{indicadores['cumplimiento']:.1f}%",
            delta=None,
        )
    
    st.divider()
    
    # --- DISTRIBUCIÓN DE RESULTADOS ---
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("📈 Distribución de Resultados")
        distribucion = indicadores["distribucion"]
        if distribucion:
            df_dist = pd.DataFrame([
                {
                    "Resultado": k,
                    "Cantidad": v["cantidad"],
                    "Porcentaje": f"{v['porcentaje']:.1f}%"
                }
                for k, v in distribucion.items()
            ])
            st.dataframe(df_dist, use_container_width=True, hide_index=True)
        else:
            st.info("No hay resultados registrados.")
    
    with col2:
        st.subheader("📊 Gráfico de Resultados")
        if distribucion:
            import plotly.express as px
            df_plot = pd.DataFrame([
                {"Resultado": k, "Cantidad": v["cantidad"]}
                for k, v in distribucion.items()
            ])
            fig = px.pie(
                df_plot, 
                values="Cantidad", 
                names="Resultado",
                title="Distribución de Visitas",
                color_discrete_sequence=px.colors.qualitative.Set2,
            )
            fig.update_layout(height=350, margin=dict(l=20, r=20, t=50, b=20))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No hay datos para graficar.")
    
    st.divider()
    
    # --- DETALLE DE VISITAS ---
    st.subheader("📋 Detalle de Visitas")
    df_detalle = preparar_detalle_visitas(df_filtrado)
    if not df_detalle.empty:
        st.dataframe(df_detalle, use_container_width=True, hide_index=True)
    else:
        st.info("No hay visitas para mostrar.")
    
    st.divider()
    
    # --- BOTÓN DE DESCARGA ---
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("📥 Descargar Reporte Excel", type="primary", use_container_width=True):
            with st.spinner("Generando reporte..."):
                excel_bytes = generar_reporte_consolidado_excel(
                    df_historial=df_historial,
                    agencia=agencia,
                    fecha_desde=fecha_desde,
